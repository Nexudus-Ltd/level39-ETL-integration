#!/usr/bin/env python3
"""
Nexudus Invoices → ETL Excel Copier → S3 Versioned Upload
Handles both XLS (old Excel) and XLSX (new Excel) formats from Nexudus
"""

import os
import sys
from datetime import datetime, timedelta
import requests
from io import BytesIO
from openpyxl import load_workbook
import boto3
from botocore.exceptions import ClientError
import xlrd

# --------------------------------------------------
# CONFIG
# --------------------------------------------------
TEMPLATE_FILE = "template/ETL_MarioDemo.xlsx"
OUTPUT_DIR = "output"
DEST_SHEET_NAME = "Membership invoices"

NEXUDUS_REPORT_URL = "https://reports.nexudus.com/ReportCenter/Invoices"
NEXUDUS_TOKEN_URL = "https://spaces.nexudus.com/api/token"

# S3
S3_BUCKET = os.getenv("S3_BUCKET", "level39-etl-mario")
S3_KEY_PREFIX = "output/"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# --------------------------------------------------
# Generate timestamped filename
# --------------------------------------------------
def generate_output_filename():
    now = datetime.utcnow()
    timestamp = now.strftime("%Y-%m-%d_%H%M%S")
    filename = f"ETL_MarioDemo_{timestamp}.xlsx"
    
    return {
        "local": os.path.join(OUTPUT_DIR, filename),
        "s3": f"{S3_KEY_PREFIX}{filename}",
        "displayName": filename
    }

# --------------------------------------------------
# Calculate yesterday's date range
# --------------------------------------------------
def get_date_range():
    yesterday = datetime.utcnow() - timedelta(days=1)
    
    start_date = datetime(yesterday.year, yesterday.month, yesterday.day, 0, 0, 0)
    end_date = datetime(yesterday.year, yesterday.month, yesterday.day, 23, 59, 59)
    
    return {
        "start": start_date.isoformat(),
        "end": end_date.isoformat()
    }

# --------------------------------------------------
# Get Nexudus Access Token
# --------------------------------------------------
def get_nexudus_token():
    try:
        print("🔑 Requesting Nexudus authentication token...")
        
        response = requests.post(
            NEXUDUS_TOKEN_URL,
            data={
                "grant_type": "password",
                "username": os.getenv("NEXUDUS_USERNAME"),
                "password": os.getenv("NEXUDUS_PASSWORD")
            },
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Accept": "application/json"
            },
            timeout=10
        )
        
        response.raise_for_status()
        
        data = response.json()
        if "access_token" not in data:
            raise Exception("No access token received from Nexudus")
        
        print("✔ Nexudus token obtained successfully")
        return data["access_token"]
        
    except requests.exceptions.RequestException as err:
        if "401" in str(err):
            raise Exception("Nexudus authentication failed - invalid credentials")
        elif "timeout" in str(err).lower():
            raise Exception("Cannot reach Nexudus API - network issue or service unavailable")
        raise Exception(f"Nexudus token request failed: {str(err)}")

# --------------------------------------------------
# Download Nexudus report
# --------------------------------------------------
def download_report(token):
    try:
        date_range = get_date_range()
        start = date_range["start"].split(".")[0]
        end = date_range["end"].split(".")[0]
        
        print(f"📅 Fetching invoices from {start} to {end}")
        
        response = requests.get(
            NEXUDUS_REPORT_URL,
            params={
                "businessId": os.getenv("BUSINESS_ID"),
                "reportName": "Invoices/InvoicesAccount",
                "start": start,
                "end": end,
                "format": "Excel",
                "portrait": "false",
                "rnd": datetime.utcnow().isoformat()
            },
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
            timeout=30
        )
        
        response.raise_for_status()
        
        if len(response.content) == 0:
            raise Exception("Nexudus report returned empty data")
        
        print(f"✔ Report downloaded successfully ({len(response.content)} bytes)")
        
        # Check file format
        file_sig = response.content[:8]
        if file_sig.startswith(b'\xd0\xcf\x11\xe0'):
            print(f"  ├─ ✔ File is OLE2/XLS format (old Excel)")
            file_type = "xls"
        elif file_sig.startswith(b'PK'):
            print(f"  ├─ ✔ File is XLSX format (new Excel)")
            file_type = "xlsx"
        else:
            raise Exception(f"Unknown file format. First bytes: {file_sig}")
        
        return BytesIO(response.content), file_type
        
    except requests.exceptions.RequestException as err:
        if "404" in str(err):
            raise Exception("Nexudus report endpoint not found - check API URL")
        elif "403" in str(err):
            raise Exception("Access denied to Nexudus report - check business ID and permissions")
        elif "timeout" in str(err).lower():
            raise Exception("Nexudus report request timed out")
        raise Exception(f"Failed to download Nexudus report: {str(err)}")

# --------------------------------------------------
# Convert XLS to XLSX if needed
# --------------------------------------------------
def convert_xls_to_xlsx(file_buffer):
    """Convert XLS (OLE2) file to XLSX format"""
    try:
        print("  ├─ Converting XLS to XLSX format...")
        
        # Read the XLS file
        file_buffer.seek(0)
        xls_workbook = xlrd.open_workbook(file_contents=file_buffer.read())
        
        print(f"  ├─ XLS has {len(xls_workbook.sheet_names())} sheets: {', '.join(xls_workbook.sheet_names())}")
        
        # Create a new XLSX workbook
        xlsx_workbook = load_workbook()
        xlsx_workbook.remove(xlsx_workbook.active)  # Remove default sheet
        
        # Copy each sheet from XLS to XLSX
        for sheet_idx, sheet_name in enumerate(xls_workbook.sheet_names()):
            xls_sheet = xls_workbook.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_workbook.create_sheet(title=sheet_name)
            
            print(f"  ├─ Copying sheet '{sheet_name}' ({xls_sheet.nrows} rows, {xls_sheet.ncols} cols)")
            
            # Copy all cells
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None and cell_value != '':
                        xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        print(f"  └─ Conversion complete")
        
        # Save to buffer
        output_buffer = BytesIO()
        xlsx_workbook.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer
        
    except Exception as err:
        raise Exception(f"Failed to convert XLS to XLSX: {str(err)}")

# --------------------------------------------------
# Replace Membership invoices sheet with Nexudus data
# --------------------------------------------------
def update_template_with_nexudus_sheet(report_buffer, file_type, output_file):
    try:
        print("📝 Processing Excel data...")
        
        # Convert if XLS format
        if file_type == "xls":
            report_buffer = convert_xls_to_xlsx(report_buffer)
        
        # Load the Nexudus report
        print("  ├─ Loading Nexudus report...")
        report_buffer.seek(0)
        report_wb = load_workbook(report_buffer)
        
        if not report_wb.sheetnames:
            raise Exception("Nexudus workbook contains no sheets")
        
        print(f"  ├─ Available sheets: {', '.join(report_wb.sheetnames)}")
        
        # Get the first sheet from Nexudus
        source_sheet_name = report_wb.sheetnames[0]
        source_sheet = report_wb[source_sheet_name]
        
        print(f"  ├─ Using sheet: '{source_sheet_name}'")
        print(f"  ├─ Sheet has {source_sheet.max_row} rows and {source_sheet.max_column} columns")
        
        # Load template
        print("  ├─ Loading template...")
        
        if not os.path.exists(TEMPLATE_FILE):
            raise Exception(f"Template file not found: {TEMPLATE_FILE}")
        
        file_size = os.path.getsize(TEMPLATE_FILE)
        print(f"  ├─ Template file size: {file_size} bytes")
        
        template_wb = load_workbook(TEMPLATE_FILE)
        print(f"  ├─ Template sheets: {', '.join(template_wb.sheetnames)}")
        
        if DEST_SHEET_NAME not in template_wb.sheetnames:
            raise Exception(f"Sheet '{DEST_SHEET_NAME}' not found in template")
        
        print(f"  ├─ Removing old '{DEST_SHEET_NAME}' sheet...")
        template_wb.remove(template_wb[DEST_SHEET_NAME])
        
        print(f"  ├─ Copying Nexudus sheet into template...")
        source_sheet_copy = template_wb.copy_worksheet(source_sheet)
        source_sheet_copy.title = DEST_SHEET_NAME
        
        print(f"  ├─ Template sheets now: {', '.join(template_wb.sheetnames)}")
        print(f"  └─ Saving to file...")
        
        # Save to output file
        template_wb.save(output_file)
        print(f"✔ Local ETL file saved: {output_file}")
        
    except Exception as err:
        raise Exception(f"Failed to process Excel template: {str(err)}")

# --------------------------------------------------
# Upload to S3
# --------------------------------------------------
def upload_to_s3(local_file, s3_key, display_name):
    try:
        print("☁️  Uploading to S3...")
        
        s3_client = boto3.client(
            "s3",
            region_name=os.getenv("AWS_REGION", "us-east-1"),
            aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
            aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY")
        )
        
        with open(local_file, "rb") as f:
            response = s3_client.put_object(
                Bucket=S3_BUCKET,
                Key=s3_key,
                Body=f.read()
            )
        
        print(f"✔ File uploaded to S3: s3://{S3_BUCKET}/{s3_key}")
        
        if "VersionId" in response:
            print(f"📦 Version ID: {response['VersionId']}")
        else:
            print("⚠️  Note: S3 versioning may not be enabled on this bucket")
            
    except ClientError as err:
        if err.response['Error']['Code'] == 'NoSuchBucket':
            raise Exception(f"S3 bucket not found: {S3_BUCKET}")
        elif err.response['Error']['Code'] == 'AccessDenied':
            raise Exception("Access denied to S3 bucket - check AWS credentials and permissions")
        raise Exception(f"Failed to upload to S3: {str(err)}")

# --------------------------------------------------
# Main ETL Job
# --------------------------------------------------
def run_etl():
    try:
        print("=" * 50)
        print("🚀 Starting Nexudus ETL Job")
        print("=" * 50)
        print(f"⏰ Run time: {datetime.utcnow().isoformat()}")
        print()
        
        file_names = generate_output_filename()
        print(f"📁 Output file: {file_names['displayName']}")
        print()
        
        # Validate environment variables
        required_vars = [
            "NEXUDUS_USERNAME",
            "NEXUDUS_PASSWORD",
            "BUSINESS_ID",
            "AWS_ACCESS_KEY_ID",
            "AWS_SECRET_ACCESS_KEY"
        ]
        
        missing_vars = [var for var in required_vars if not os.getenv(var)]
        if missing_vars:
            raise Exception(f"Missing required environment variables: {', '.join(missing_vars)}")
        
        # Execute ETL pipeline
        token = get_nexudus_token()
        report_buffer, file_type = download_report(token)
        update_template_with_nexudus_sheet(report_buffer, file_type, file_names["local"])
        upload_to_s3(file_names["local"], file_names["s3"], file_names["displayName"])
        
        print()
        print("=" * 50)
        print("✅ ETL Job Completed Successfully")
        print("=" * 50)
        
    except Exception as err:
        print()
        print("=" * 50)
        print("❌ ETL Job Failed")
        print("=" * 50)
        print(f"⏰ Error time: {datetime.utcnow().isoformat()}")
        print(f"📌 Error: {str(err)}")
        print("=" * 50)
        
        if os.getenv("DEBUG") == "true":
            import traceback
            print("\nFull stack trace:")
            traceback.print_exc()
        
        sys.exit(1)

# --------------------------------------------------
# Run ETL
# --------------------------------------------------
if __name__ == "__main__":
    run_etl()